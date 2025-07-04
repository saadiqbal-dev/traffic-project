"""
Automated Iterative Stress Test Script
=====================================
This script automates the iterative stress test process (Option Z from main menu).

Features:
- Automated 30 iterations with 20 vehicles per iteration
- Point A to Point B selection using the same notable locations
- Exact same Excel export as the manual process
- No manual intervention required during execution
- Uses existing system functions for consistency

Usage:
    python automated_iterative_stress_test.py
"""

import os
import sys
import time
import random
import numpy as np
from typing import List, Dict, Optional, Tuple

# Import all required modules from the existing system
from models import (
    Vehicle, load_london_network, generate_initial_congestion, 
    calculate_shortest_path, update_vehicle_counts_for_path,
    create_evenly_distributed_notable_locations
)
from routing import calculate_all_routes
from congestion import (
    apply_consistent_congestion_scenario, update_congestion_based_on_vehicles
)
from vehicle_management import calculate_overall_congestion_metrics
from stress_testing import export_iterative_stress_test_to_excel
from analysis import open_excel_file
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.marker import DataPoint


class AutomatedIterativeStressTest:
    """
    Automated version of the iterative stress test from the main menu (Option Z).
    Runs 30 iterations with 20 vehicles per iteration automatically.
    """
    
    def __init__(self):
        self.G = None
        self.vehicles = []
        self.congestion_data = {}
        self.original_congestion = {}
        self.notable_locations = {}
        self.scenario = "Normal"  # Default scenario
        
    def initialize_system(self) -> bool:
        """
        Initialize the traffic simulation system.
        
        Returns:
            True if initialization successful, False otherwise
        """
        print("=" * 80)
        print("AUTOMATED ITERATIVE STRESS TEST")
        print("=" * 80)
        print("Initializing London traffic simulation system...")
        
        try:
            # Load the London network (same as main system)
            print("Loading London street network...")
            self.G = load_london_network()
            print(f"✓ Graph loaded: {self.G.number_of_nodes()} nodes, {self.G.number_of_edges()} edges")
            
            # Generate initial congestion
            print("Generating initial congestion...")
            self.congestion_data = generate_initial_congestion(self.G)
            self.original_congestion = self.congestion_data.copy()
            print("✓ Initial congestion generated")
            
            # Create notable locations (same as main system)
            print("Creating notable locations...")
            self.notable_locations = create_evenly_distributed_notable_locations(self.G)
            print("✓ Notable locations created:")
            for i, (name, node) in enumerate(self.notable_locations.items(), 1):
                print(f"  {i}. {name}: Node {node}")
            
            # Apply default scenario
            print(f"Applying {self.scenario} congestion scenario...")
            self.congestion_data, _ = apply_consistent_congestion_scenario(
                self.G, self.congestion_data, self.scenario, self.original_congestion
            )
            print(f"✓ {self.scenario} scenario applied")
            
            return True
            
        except Exception as e:
            print(f"✗ Error during initialization: {e}")
            return False
    
    def select_route_points(self) -> Tuple[Optional[int], Optional[int]]:
        """
        Let user select Point A and Point B from notable locations.
        
        Returns:
            Tuple of (source_node, destination_node) or (None, None) if invalid
        """
        print("\n" + "=" * 60)
        print("ROUTE SELECTION")
        print("=" * 60)
        print("Select Point A (Source) and Point B (Destination) for the stress test.")
        print("Available notable locations:")
        
        location_list = list(self.notable_locations.items())
        for i, (name, node) in enumerate(location_list, 1):
            print(f"  {i}. {name} (Node: {node})")
        
        try:
            # Select Point A (Source)
            print("\nSelect Point A (Source):")
            source_choice = int(input("Enter location number (1-{}): ".format(len(location_list))))
            
            if not (1 <= source_choice <= len(location_list)):
                print("Invalid source selection!")
                return None, None
            
            source_name, source_node = location_list[source_choice - 1]
            print(f"✓ Point A selected: {source_name} (Node: {source_node})")
            
            # Select Point B (Destination)
            print("\nSelect Point B (Destination):")
            dest_choice = int(input("Enter location number (1-{}): ".format(len(location_list))))
            
            if not (1 <= dest_choice <= len(location_list)):
                print("Invalid destination selection!")
                return None, None
            
            if dest_choice == source_choice:
                print("Source and destination cannot be the same!")
                return None, None
            
            dest_name, dest_node = location_list[dest_choice - 1]
            print(f"✓ Point B selected: {dest_name} (Node: {dest_node})")
            
            print(f"\nRoute confirmed: {source_name} → {dest_name}")
            return source_node, dest_node
            
        except ValueError:
            print("Invalid input! Please enter a number.")
            return None, None
        except KeyboardInterrupt:
            print("\nOperation cancelled by user.")
            return None, None
    
    def create_test_vehicle(self, source: int, destination: int) -> Optional[Vehicle]:
        """
        Create the main test vehicle for the stress test.
        
        Args:
            source: Source node
            destination: Destination node
            
        Returns:
            Created Vehicle object or None if failed
        """
        print(f"\nCreating test vehicle from {source} to {destination}...")
        
        # Calculate initial path
        path = calculate_shortest_path(self.G, source, destination)
        
        if not path:
            print("✗ Could not create test vehicle: No valid path found!")
            return None
        
        # Create the vehicle
        vehicle = Vehicle(1, source, destination, path)
        
        # Calculate routes using all algorithms
        print("Calculating initial routes using all algorithms...")
        vehicle = calculate_all_routes(self.G, vehicle, self.congestion_data)
        
        # Add to vehicles list
        self.vehicles.append(vehicle)
        
        # Update vehicle count on the path
        update_vehicle_counts_for_path(self.G, path, 1)
        
        print(f"✓ Test vehicle created (ID: {vehicle.id})")
        print(f"  Path length: {len(path)} nodes")
        
        if vehicle.travel_times:
            for algo, time_val in vehicle.travel_times.items():
                print(f"  {algo} travel time: {time_val:.2f}s")
        
        return vehicle
    
    def add_stress_vehicles_iteration(self, original_vehicle: Vehicle, count: int = 20) -> Tuple[int, List[Vehicle]]:
        """
        Add stress test vehicles for one iteration (same logic as existing system).
        
        Args:
            original_vehicle: The main test vehicle
            count: Number of vehicles to add (default: 20)
            
        Returns:
            Tuple of (vehicles_added, list_of_added_vehicles)
        """
        if not original_vehicle or not original_vehicle.path:
            return 0, []
        
        # Get nodes along the original path
        path_nodes = original_vehicle.path.copy()
        
        # Ensure we have enough nodes for random placement
        if len(path_nodes) < 4:
            # Extend with neighbors
            extended_nodes = set(path_nodes)
            for node in path_nodes:
                if node in self.G:
                    extended_nodes.update(list(self.G.neighbors(node)))
            path_nodes = list(extended_nodes)
        
        added_count = 0
        added_vehicles = []
        
        # Add vehicles with sources and destinations along the path
        for _ in range(count):
            if len(path_nodes) >= 2:
                # Select random source and destination from path nodes
                source = random.choice(path_nodes)
                # Ensure destination is different
                dest_candidates = [n for n in path_nodes if n != source]
                if dest_candidates:
                    destination = random.choice(dest_candidates)
                    
                    # Calculate path for this stress test vehicle
                    path = calculate_shortest_path(self.G, source, destination)
                    
                    if path:
                        vehicle_id = len(self.vehicles) + 1
                        vehicle = Vehicle(vehicle_id, source, destination, path)
                        self.vehicles.append(vehicle)
                        update_vehicle_counts_for_path(self.G, path, 1)
                        added_count += 1
                        added_vehicles.append(vehicle)
        
        return added_count, added_vehicles
    
    def export_enhanced_excel_report(self, original_vehicle: Vehicle, iterations_data: List[Dict],
                                   initial_path: List[int], total_added_vehicles: List[Vehicle],
                                   source: int, destination: int) -> str:
        """
        Export enhanced Excel report with line charts for travel time trends.
        
        Args:
            original_vehicle: The main test vehicle
            iterations_data: List of iteration data
            initial_path: Initial path before stress testing
            total_added_vehicles: List of vehicles added during stress test
            source: Source node
            destination: Destination node
            
        Returns:
            Path to the created Excel file
        """
        # Create output directory
        stress_test_dir = os.path.join('london_simulation', 'iterative_stress_test')
        os.makedirs(stress_test_dir, exist_ok=True)
        
        # Create a timestamp for the filename
        timestamp = int(time.time())
        excel_file = os.path.join(stress_test_dir, f"automated_stress_test_{source}_to_{destination}_{timestamp}.xlsx")
        
        # Create a new workbook
        wb = Workbook()
        
        # Remove default worksheet
        ws0 = wb.active
        wb.remove(ws0)
        
        # 1. Summary sheet
        ws_summary = wb.create_sheet("Summary")
        
        # Get location names for display
        location_names = list(self.notable_locations.keys())
        source_name = "Unknown"
        dest_name = "Unknown"
        
        for name, node in self.notable_locations.items():
            if node == source:
                source_name = name
            if node == destination:
                dest_name = name
        
        # Summary information
        summary_data = [
            ["Automated Iterative Stress Test Analysis"],
            [""],
            ["Generated on:", time.strftime("%Y-%m-%d %H:%M:%S")],
            ["Scenario:", self.scenario],
            ["Test Vehicle ID:", original_vehicle.id],
            ["Source:", f"{source_name} (Node: {source})"],
            ["Destination:", f"{dest_name} (Node: {destination})"],
            ["Total iterations:", len(iterations_data) - 1],  # Subtract 1 for initial state
            ["Total vehicles added:", len(total_added_vehicles)],
            ["Vehicles per iteration:", 20],
            [""],
            ["Final Path Comparison:"],
            ["Initial path length:", len(initial_path)],
            ["Final path length:", len(original_vehicle.path)],
        ]
        
        # Calculate path difference
        if initial_path != original_vehicle.path:
            diff_count = sum(1 for x, y in zip(initial_path, original_vehicle.path) if x != y)
            if len(initial_path) != len(original_vehicle.path):
                diff_count += abs(len(initial_path) - len(original_vehicle.path))
                
            diff_percentage = (diff_count / max(len(initial_path), len(original_vehicle.path))) * 100
            summary_data.append(["Path difference:", f"{diff_percentage:.1f}% of nodes are different"])
            
            # Compare initial and final travel times if available
            if iterations_data and iterations_data[0]['algorithm_results'] and iterations_data[-1]['algorithm_results']:
                for algo_name in iterations_data[0]['algorithm_results']:
                    if algo_name in iterations_data[-1]['algorithm_results']:
                        initial_time = iterations_data[0]['algorithm_results'][algo_name]['travel_time']
                        final_time = iterations_data[-1]['algorithm_results'][algo_name]['travel_time']
                        change_pct = ((final_time - initial_time) / initial_time) * 100 if initial_time > 0 else 0
                        
                        summary_data.append([f"{algo_name} travel time change:", 
                                           f"{initial_time:.2f}s → {final_time:.2f}s ({change_pct:.1f}%)"])
        else:
            summary_data.append(["Path difference:", "No change in path"])
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Format summary sheet
        from openpyxl.styles import Font, Alignment
        ws_summary.merge_cells('A1:B1')
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A1'].alignment = Alignment(horizontal='center')
        
        # 2. Chart Data sheet - Prepare data for charts
        ws_chart_data = wb.create_sheet("Chart Data")
        
        # Create headers
        ws_chart_data.append(["Iteration", "A* Travel Time", "Shortest Path Travel Time", "Shortest Path Congestion Aware Travel Time"])
        
        # Extract data for charts
        algorithms = ["A*", "Shortest Path", "Shortest Path Congestion Aware"]
        chart_data = []
        
        for iter_data in iterations_data:
            row = [iter_data['iteration']]
            for algo in algorithms:
                if algo in iter_data['algorithm_results']:
                    travel_time = iter_data['algorithm_results'][algo]['travel_time']
                    row.append(travel_time)
                else:
                    row.append(0)  # Default value if algorithm not found
            chart_data.append(row)
            ws_chart_data.append(row)
        
        # 3. Create individual chart sheets for each algorithm
        chart_colors = ["0066CC", "FF6600", "00AA00"]  # Blue, Orange, Green
        
        for i, algorithm in enumerate(algorithms):
            # Create a new sheet for this algorithm's chart
            ws_chart = wb.create_sheet(f"{algorithm.replace('*', 'Star').replace(' ', '_')}_Chart")
            
            # Add title and data
            ws_chart.append([f"{algorithm} Travel Time Trend"])
            ws_chart.append([])
            ws_chart.append(["Iteration", "Travel Time (seconds)"])
            
            # Add data for this algorithm
            for iter_data in iterations_data:
                iteration = iter_data['iteration']
                if algorithm in iter_data['algorithm_results']:
                    travel_time = iter_data['algorithm_results'][algorithm]['travel_time']
                    ws_chart.append([iteration, travel_time])
                else:
                    ws_chart.append([iteration, 0])
            
            # Create line chart
            chart = LineChart()
            chart.title = f"{algorithm} Travel Time vs Iteration"
            chart.style = 10
            chart.y_axis.title = 'Travel Time (seconds)'
            chart.x_axis.title = 'Iteration Number'
            chart.width = 15
            chart.height = 10
            
            # Define data range (skip header rows)
            data_rows = len(iterations_data)
            data = Reference(ws_chart, min_col=2, min_row=4, max_row=3 + data_rows, max_col=2)
            cats = Reference(ws_chart, min_col=1, min_row=4, max_row=3 + data_rows, max_col=1)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            
            # Customize the line
            series = chart.series[0]
            series.graphicalProperties.line.solidFill = chart_colors[i]
            series.graphicalProperties.line.width = 25000  # Line width
            series.marker.symbol = "circle"
            series.marker.size = 5
            
            # Add chart to worksheet
            ws_chart.add_chart(chart, "D1")
            
            # Format the title
            ws_chart['A1'].font = Font(bold=True, size=14)
            ws_chart['A1'].alignment = Alignment(horizontal='center')
            ws_chart.merge_cells('A1:F1')
        
        # 4. Combined Chart sheet
        ws_combined = wb.create_sheet("Combined_Trends_Chart")
        
        # Add title and reference to chart data
        ws_combined.append(["Travel Time Trends - All Algorithms"])
        ws_combined.append([])
        ws_combined.append(["Data source: Chart Data sheet"])
        ws_combined.append([])
        
        # Create combined line chart
        combined_chart = LineChart()
        combined_chart.title = "Travel Time Trends - All Algorithms"
        combined_chart.style = 10
        combined_chart.y_axis.title = 'Travel Time (seconds)'
        combined_chart.x_axis.title = 'Iteration Number'
        combined_chart.width = 20
        combined_chart.height = 12
        
        # Reference data from Chart Data sheet
        data_rows = len(iterations_data)
        
        # Add data for each algorithm
        for i, algorithm in enumerate(algorithms):
            col_idx = i + 2  # Column B, C, D (2, 3, 4)
            data = Reference(ws_chart_data, min_col=col_idx, min_row=2, max_row=1 + data_rows, max_col=col_idx)
            combined_chart.add_data(data, titles_from_data=False)
            
            # Customize each series
            series = combined_chart.series[i]
            series.graphicalProperties.line.solidFill = chart_colors[i]
            series.graphicalProperties.line.width = 25000
            series.marker.symbol = "circle"
            series.marker.size = 5
        
        # Set categories (iterations)
        cats = Reference(ws_chart_data, min_col=1, min_row=2, max_row=1 + data_rows, max_col=1)
        combined_chart.set_categories(cats)
        
        # Add legend
        combined_chart.legend.position = 'r'  # Right side
        
        # Manually set legend labels using SeriesLabel
        from openpyxl.chart.series import SeriesLabel
        for i, algorithm in enumerate(algorithms):
            series_label = SeriesLabel(v=algorithm)
            combined_chart.series[i].tx = series_label
        
        # Add chart to worksheet
        ws_combined.add_chart(combined_chart, "A6")
        
        # Format the title
        ws_combined['A1'].font = Font(bold=True, size=16)
        ws_combined['A1'].alignment = Alignment(horizontal='center')
        ws_combined.merge_cells('A1:H1')
        
        # 5. Raw Data sheet (same as original)
        ws_raw = wb.create_sheet("Raw Data")
        
        # Create table with data from all iterations
        ws_raw.append(["Automated Stress Test - Raw Data"])
        ws_raw.append([])
        
        # Create detailed data rows
        iteration_rows = []
        for iter_data in iterations_data:
            # Get algorithm data
            for algo_name, algo_results in iter_data['algorithm_results'].items():
                iteration_rows.append({
                    'Iteration': iter_data['iteration'],
                    'Vehicles Added': iter_data['vehicles_added'],
                    'Total Vehicles': iter_data['total_vehicles'],
                    'Mean Congestion': iter_data['mean_congestion'],
                    'Algorithm': algo_name,
                    'Travel Time (s)': algo_results['travel_time'],
                    'Path Length': algo_results['path_length'],
                    'Computation Time (s)': algo_results['computation_time']
                })
        
        # Write headers
        if iteration_rows:
            headers = list(iteration_rows[0].keys())
            ws_raw.append(headers)
            
            # Write data
            for row_data in iteration_rows:
                ws_raw.append([row_data[header] for header in headers])
        
        # 6. Added Vehicles sheet
        ws_vehicles = wb.create_sheet("Added Vehicles")
        
        ws_vehicles.append(["Vehicles Added During Stress Test"])
        ws_vehicles.append([])
        ws_vehicles.append(["Vehicle ID", "Source", "Destination", "Path Length"])
        
        for v in total_added_vehicles:
            ws_vehicles.append([v.id, v.source, v.destination, len(v.path) if v.path else 0])
        
        # Save the workbook
        wb.save(excel_file)
        print(f"Enhanced Excel report with charts exported to {excel_file}")
        
        return excel_file
    
    def run_automated_stress_test(self, source: int, destination: int,
                                 iterations: int = 30, vehicles_per_iteration: int = 20) -> Optional[str]:
        """
        Run the automated iterative stress test.
        
        Args:
            source: Source node
            destination: Destination node
            iterations: Number of iterations (default: 30)
            vehicles_per_iteration: Vehicles to add per iteration (default: 20)
            
        Returns:
            Path to generated Excel file or None if failed
        """
        print("\n" + "=" * 80)
        print("STARTING AUTOMATED ITERATIVE STRESS TEST")
        print("=" * 80)
        print(f"Route: Node {source} → Node {destination}")
        print(f"Iterations: {iterations}")
        print(f"Vehicles per iteration: {vehicles_per_iteration}")
        print(f"Total vehicles to be added: {iterations * vehicles_per_iteration}")
        print()
        
        # Create the main test vehicle
        original_vehicle = self.create_test_vehicle(source, destination)
        if not original_vehicle:
            return None
        
        # Store initial state
        initial_path = original_vehicle.path.copy()
        initial_travel_times = original_vehicle.travel_times.copy() if original_vehicle.travel_times else {}
        
        # Track iterations data (same format as existing system)
        iterations_data = []
        iteration_count = 0
        total_added_vehicles = []
        
        # Initial congestion analysis
        initial_congestion_stats = calculate_overall_congestion_metrics(self.G, self.congestion_data)
        
        print("Initial route analysis:")
        if original_vehicle.travel_times:
            for algo, time_val in original_vehicle.travel_times.items():
                print(f"  {algo}: {time_val:.2f}s")
        
        # Store initial state in iterations data
        algorithm_results = {}
        for algo_name, travel_time in original_vehicle.travel_times.items():
            path = original_vehicle.paths.get(algo_name, [])
            algorithm_results[algo_name] = {
                'travel_time': travel_time,
                'path_length': len(path),
                'computation_time': original_vehicle.computation_times.get(algo_name, 0)
            }
        
        iterations_data.append({
            'iteration': iteration_count,
            'vehicles_added': 0,
            'total_vehicles': len(self.vehicles),
            'mean_congestion': initial_congestion_stats['mean'],
            'algorithm_results': algorithm_results.copy(),
            'timestamp': time.time()
        })
        
        print(f"\nStarting automated iterations...")
        print("Progress: [" + " " * 50 + "] 0%", end="\r")
        
        # Main automation loop
        for iteration in range(1, iterations + 1):
            iteration_count = iteration
            
            # Update progress bar
            progress = int((iteration / iterations) * 50)
            progress_bar = "[" + "=" * progress + " " * (50 - progress) + "]"
            percentage = int((iteration / iterations) * 100)
            print(f"Progress: {progress_bar} {percentage}% (Iteration {iteration}/{iterations})", end="\r")
            
            # Add vehicles for this iteration
            added_count, added_vehicles = self.add_stress_vehicles_iteration(
                original_vehicle, vehicles_per_iteration
            )
            total_added_vehicles.extend(added_vehicles)
            
            # Update congestion based on new vehicles
            update_congestion_based_on_vehicles(self.G, self.congestion_data, self.original_congestion)
            
            # Recalculate routes for the original vehicle
            calculate_all_routes(self.G, original_vehicle, self.congestion_data)
            
            # Calculate congestion metrics after this iteration
            current_congestion_stats = calculate_overall_congestion_metrics(self.G, self.congestion_data)
            
            # Store this iteration's results
            algorithm_results = {}
            for algo_name, travel_time in original_vehicle.travel_times.items():
                path = original_vehicle.paths.get(algo_name, [])
                algorithm_results[algo_name] = {
                    'travel_time': travel_time,
                    'path_length': len(path),
                    'computation_time': original_vehicle.computation_times.get(algo_name, 0)
                }
            
            iterations_data.append({
                'iteration': iteration_count,
                'vehicles_added': added_count,
                'total_vehicles': len(self.vehicles),
                'mean_congestion': current_congestion_stats['mean'],
                'algorithm_results': algorithm_results.copy(),
                'timestamp': time.time()
            })
        
        print(f"\nProgress: [" + "=" * 50 + "] 100% (Complete!)")
        
        # Final analysis
        print("\n" + "=" * 80)
        print("STRESS TEST COMPLETED")
        print("=" * 80)
        print(f"Total iterations completed: {iterations}")
        print(f"Total vehicles added: {len(total_added_vehicles)}")
        print(f"Final total vehicles: {len(self.vehicles)}")
        
        # Compare initial vs final
        print("\nRoute comparison (Initial vs Final):")
        if original_vehicle.travel_times and initial_travel_times:
            for algo in initial_travel_times:
                if algo in original_vehicle.travel_times:
                    initial_time = initial_travel_times[algo]
                    final_time = original_vehicle.travel_times[algo]
                    change_pct = ((final_time - initial_time) / initial_time) * 100 if initial_time > 0 else 0
                    print(f"  {algo}: {initial_time:.2f}s → {final_time:.2f}s ({change_pct:+.1f}%)")
        
        # Path change analysis
        if initial_path != original_vehicle.path:
            diff_count = sum(1 for x, y in zip(initial_path, original_vehicle.path) if x != y)
            if len(initial_path) != len(original_vehicle.path):
                diff_count += abs(len(initial_path) - len(original_vehicle.path))
            diff_percentage = (diff_count / max(len(initial_path), len(original_vehicle.path))) * 100
            print(f"\nPath changed: {diff_percentage:.1f}% of nodes are different")
            print(f"  Initial path length: {len(initial_path)} nodes")
            print(f"  Final path length: {len(original_vehicle.path)} nodes")
        else:
            print("\nPath remained unchanged despite increased congestion")
        
        # Export to Excel with enhanced charts
        print("\nGenerating Excel report with trend charts...")
        excel_file = self.export_enhanced_excel_report(
            original_vehicle, iterations_data, initial_path, 
            total_added_vehicles, source, destination
        )
        
        print(f"✓ Excel report with charts generated: {excel_file}")
        return excel_file
    
    def run(self) -> bool:
        """
        Main entry point for the automated stress test.
        
        Returns:
            True if successful, False otherwise
        """
        try:
            # Initialize the system
            if not self.initialize_system():
                return False
            
            # Get route selection from user
            source, destination = self.select_route_points()
            if source is None or destination is None:
                print("Invalid route selection. Exiting.")
                return False
            
            # Confirm before starting
            print("\n" + "=" * 60)
            print("CONFIRMATION")
            print("=" * 60)
            print("The automated stress test will:")
            print("• Run 30 iterations automatically")
            print("• Add 20 vehicles per iteration")
            print("• Generate comprehensive Excel report")
            print("• Take approximately 5-10 minutes to complete")
            print()
            
            confirm = input("Proceed with automated stress test? (y/n): ").strip().lower()
            if confirm != 'y':
                print("Stress test cancelled.")
                return False
            
            # Run the automated stress test
            excel_file = self.run_automated_stress_test(source, destination)
            
            if excel_file:
                print("\n" + "=" * 80)
                print("AUTOMATED STRESS TEST COMPLETED SUCCESSFULLY!")
                print("=" * 80)
                print(f"Excel report saved to: {excel_file}")
                
                # Ask if user wants to open the Excel file
                open_file = input("\nOpen the Excel report now? (y/n): ").strip().lower()
                if open_file == 'y':
                    try:
                        open_excel_file(excel_file)
                        print("✓ Excel file opened")
                    except Exception as e:
                        print(f"Could not open Excel file automatically: {e}")
                        print(f"Please open manually: {excel_file}")
                
                return True
            else:
                print("✗ Stress test failed to generate Excel report")
                return False
                
        except KeyboardInterrupt:
            print("\n\nStress test interrupted by user. Exiting...")
            return False
        except Exception as e:
            print(f"\n✗ Error during stress test: {e}")
            import traceback
            traceback.print_exc()
            return False


def main():
    """
    Main function to run the automated iterative stress test.
    """
    print("Automated Iterative Stress Test")
    print("This script automates the iterative stress test process from the main menu (Option Z)")
    print()
    
    # Check dependencies
    try:
        import osmnx
        import networkx
        import matplotlib.pyplot as plt
        import pandas as pd
        import numpy as np
        import openpyxl
        print("✓ All required dependencies are available")
    except ImportError as e:
        print(f"✗ Missing required dependency: {e}")
        print("\nPlease install required packages:")
        print("pip install osmnx networkx matplotlib pandas numpy openpyxl scikit-learn")
        return False
    
    # Create and run the automated stress test
    stress_test = AutomatedIterativeStressTest()
    success = stress_test.run()
    
    if success:
        print("\nAutomated stress test completed successfully!")
        print("You can now use the generated Excel file for analysis and graph creation.")
    else:
        print("\nAutomated stress test failed or was cancelled.")
    
    return success


if __name__ == "__main__":
    main()
