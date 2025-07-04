"""
Stress testing module for the traffic simulation project.
Handles iterative stress tests and performance analysis.
"""

import time
import random
import numpy as np
import pandas as pd
import os
from typing import List, Dict, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, Series

from models import Vehicle, calculate_shortest_path, update_vehicle_counts_for_path
from routing import calculate_all_routes
from congestion import update_congestion_based_on_vehicles
from vehicle_management import calculate_overall_congestion_metrics
from visualization import enhanced_visualize_congestion_map, debug_congestion_flow, debug_travel_time_calculation, quick_congestion_check
from analysis import print_algorithm_comparison_table, simple_algorithm_comparison_table, open_excel_file


def debug_stress_test_step(G, vehicles: List[Vehicle], original_vehicle: Vehicle, 
                          congestion_data: Dict[str, float], step_name: str) -> None:
    """
    Debug helper for stress test iterations.
    
    Args:
        G: NetworkX graph
        vehicles: List of all vehicles
        original_vehicle: Vehicle being stress tested
        congestion_data: Current congestion data
        step_name: Name of the debug step
    """
    print(f"\n{'='*60}")
    print(f"DEBUG: {step_name}")
    print(f"{'='*60}")
    
    quick_congestion_check(G, congestion_data)
    
    if original_vehicle and original_vehicle.path:
        debug_congestion_flow(G, congestion_data, original_vehicle.path, step_name)
    
    debug_travel_time_calculation(G, congestion_data, 2)


def run_iterative_stress_test(G, vehicles: List[Vehicle], original_vehicle: Vehicle,
                             congestion_data: Dict[str, float], 
                             original_congestion: Dict[str, float], scenario: str) -> str:
    """
    Run an iterative stress test by repeatedly adding vehicles and analyzing impact.
    
    Args:
        G: Graph representing the road network
        vehicles: List of all vehicles
        original_vehicle: The vehicle to analyze during stress testing
        congestion_data: Current congestion data
        original_congestion: Original congestion data baseline
        scenario: Current scenario name
        
    Returns:
        Path to the Excel file with comprehensive results
    """
    print("\n=== Starting Iterative Stress Test ===")
    print(f"Testing impact on Vehicle {original_vehicle.id}: {original_vehicle.source} -> {original_vehicle.destination}")
    
    # Save initial state
    initial_path = original_vehicle.path.copy()
    initial_travel_times = original_vehicle.travel_times.copy() if original_vehicle.travel_times else {}
    initial_paths = {algo: path.copy() for algo, path in original_vehicle.paths.items()} if original_vehicle.paths else {}
    
    # Track iterations
    iterations = []
    iteration_count = 0
    added_vehicles_count = 0
    total_added_vehicles = []
    
    # Initial congestion analysis
    initial_congestion_stats = calculate_overall_congestion_metrics(G, congestion_data)
    
    # Create output directory for stress test
    stress_test_dir = os.path.join('london_simulation', 'iterative_stress_test')
    os.makedirs(stress_test_dir, exist_ok=True)
    
    # Show initial route
    print("\nInitial route before stress testing:")
    try:
        print_algorithm_comparison_table(G, original_vehicle, congestion_data)
    except Exception as e:
        print(f"Warning: Complex comparison failed, using simple version: {e}")
        simple_algorithm_comparison_table(G, original_vehicle, congestion_data)
        
    # Store initial state in iterations
    algorithm_results = {}
    for algo_name, travel_time in original_vehicle.travel_times.items():
        path = original_vehicle.paths.get(algo_name, [])
        algorithm_results[algo_name] = {
            'travel_time': travel_time,
            'path_length': len(path),
            'computation_time': original_vehicle.computation_times.get(algo_name, 0)
        }
    
    iterations.append({
        'iteration': iteration_count,
        'vehicles_added': 0,
        'total_vehicles': len(vehicles),
        'mean_congestion': initial_congestion_stats['mean'],
        'algorithm_results': algorithm_results.copy(),
        'timestamp': time.time()
    })
    
    # Iterative loop
    while True:
        iteration_count += 1
        print(f"\n--- Stress Test Iteration {iteration_count} ---")
        print(f"Total vehicles so far: {len(vehicles)}")
        
        # Ask how many vehicles to add in this iteration
        try:
            num_vehicles = int(input("\nHow many vehicles to add in this iteration (or enter 0 to stop stress test): "))
            if num_vehicles == 0:
                print("Stopping stress test...")
                break
                
            if num_vehicles < 0:
                print("Please enter a positive number.")
                continue
        except ValueError:
            print("Invalid input. Please enter a number.")
            continue
        
        # Get nodes along the original path
        path_nodes = original_vehicle.path.copy()
        
        # Ensure we have enough nodes for random placement
        if len(path_nodes) < 4:
            # Extend with neighbors
            extended_nodes = set(path_nodes)
            for node in path_nodes:
                extended_nodes.update(list(G.neighbors(node)))
            path_nodes = list(extended_nodes)
        
        # Add vehicles
        added_count = 0
        added_vehicles = []
        
        print(f"Adding {num_vehicles} vehicles along the route of Vehicle {original_vehicle.id}...")
        
        for _ in range(num_vehicles):
            if len(path_nodes) >= 2:
                # Select random source and destination from path nodes
                source = random.choice(path_nodes)
                # Ensure destination is different
                dest_candidates = [n for n in path_nodes if n != source]
                destination = random.choice(dest_candidates)
                
                # Calculate path for this stress test vehicle
                path = calculate_shortest_path(G, source, destination)
                
                if path:
                    vehicle_id = len(vehicles) + 1
                    vehicle = Vehicle(vehicle_id, source, destination, path)
                    vehicles.append(vehicle)
                    update_vehicle_counts_for_path(G, path, 1)
                    added_count += 1
                    added_vehicles.append(vehicle)
        
        added_vehicles_count += added_count
        total_added_vehicles.extend(added_vehicles)
        print(f"Added {added_count} vehicles in this iteration (total added: {added_vehicles_count})")
        
        # Update congestion based on the new vehicles
        update_congestion_based_on_vehicles(G, congestion_data, original_congestion)
        # DEBUG: Check if congestion is being updated
        debug_stress_test_step(G, vehicles, original_vehicle, congestion_data, f"After adding {added_count} vehicles")
        
        # Recalculate the route for the original vehicle
        print(f"Recalculating route for Vehicle {original_vehicle.id} under new congestion conditions...")
        calculate_all_routes(G, original_vehicle, congestion_data)
        
        # Calculate congestion metrics after this iteration
        current_congestion_stats = calculate_overall_congestion_metrics(G, congestion_data)
        
        # Display updated map
        enhanced_visualize_congestion_map(G, congestion_data, vehicles, scenario, None)
        
        # Print algorithm comparison table
        print("\nRoute comparison after this iteration:")
        try:
            print_algorithm_comparison_table(G, original_vehicle, congestion_data)
        except Exception as e:
            print(f"Warning: Complex comparison failed, using simple version: {e}")
            simple_algorithm_comparison_table(G, original_vehicle, congestion_data)
        
        # Store this iteration's results
        algorithm_results = {}
        for algo_name, travel_time in original_vehicle.travel_times.items():
            path = original_vehicle.paths.get(algo_name, [])
            algorithm_results[algo_name] = {
                'travel_time': travel_time,
                'path_length': len(path),
                'computation_time': original_vehicle.computation_times.get(algo_name, 0)
            }
        
        iterations.append({
            'iteration': iteration_count,
            'vehicles_added': added_count,
            'total_vehicles': len(vehicles),
            'mean_congestion': current_congestion_stats['mean'],
            'algorithm_results': algorithm_results.copy(),
            'timestamp': time.time()
        })
        
        # Ask if user wants to continue
        print("\nPress Enter to continue stress test or type 'V' to stop and view results: ", end="")
        stop_input = input().strip().upper()
        if stop_input == 'V':
            print("Stopping stress test and generating final report...")
            break
    
    # Export final analysis to Excel
    print("\nExporting comprehensive stress test analysis to Excel...")
    excel_file = export_iterative_stress_test_to_excel(G, original_vehicle, iterations, 
                                                      initial_path, congestion_data, 
                                                      original_congestion, scenario, 
                                                      total_added_vehicles)
    
    # Print final summary
    print(f"\nStress Test Summary:")
    print(f"  Total iterations: {iteration_count}")
    print(f"  Total vehicles added: {added_vehicles_count}")
    print(f"  Excel report saved to: {excel_file}")
    
    # Ask if the user wants to open the Excel file
    if input("\nOpen the stress test Excel report now? (y/n): ").strip().lower() == 'y':
        open_excel_file(excel_file)
    
    return excel_file


def export_iterative_stress_test_to_excel(G, original_vehicle: Vehicle, iterations: List[Dict],
                                         initial_path: List[int], congestion_data: Dict[str, float],
                                         original_congestion: Dict[str, float], scenario: str,
                                         added_vehicles: List[Vehicle]) -> str:
    """
    Export iterative stress test results to Excel with comparisons between iterations.
    
    Args:
        G: Graph representing the road network
        original_vehicle: The vehicle being stress tested
        iterations: List of dictionaries containing data from each iteration
        initial_path: Original path before stress testing
        congestion_data: Current congestion data
        original_congestion: Original congestion data
        scenario: Current scenario name
        added_vehicles: List of vehicles added during stress test
        
    Returns:
        Path to the created Excel file
    """
    # Create output directory
    stress_test_dir = os.path.join('london_simulation', 'iterative_stress_test')
    os.makedirs(stress_test_dir, exist_ok=True)
    
    # Create a timestamp for the filename
    timestamp = int(time.time())
    excel_file = os.path.join(stress_test_dir, f"iterative_stress_test_vehicle_{original_vehicle.id}_{timestamp}.xlsx")
    
    # Create a new workbook
    wb = Workbook()
    
    # Remove default worksheet
    ws0 = wb.active
    wb.remove(ws0)
    
    # 1. Summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    # Summary information
    summary_data = [
        ["Iterative Stress Test Analysis"],
        [""],
        ["Generated on:", time.strftime("%Y-%m-%d %H:%M:%S")],
        ["Scenario:", scenario],
        ["Test Vehicle ID:", original_vehicle.id],
        ["Source:", original_vehicle.source],
        ["Destination:", original_vehicle.destination],
        ["Total iterations:", len(iterations) - 1],  # Subtract 1 for initial state
        ["Total vehicles added:", sum(iter_data['vehicles_added'] for iter_data in iterations)],
        [""],
        ["Final Path Comparison:"],
        ["Initial path length:", len(initial_path)],
        ["Final path length:", len(original_vehicle.path)],
    ]
    
    # Calculate path difference
    if initial_path != original_vehicle.path:
        diff_count = sum(1 for x, y in zip(initial_path, original_vehicle.path) if x != y)
        if len(initial_path) != len(original_vehicle.path):
            diff_count += abs(len(initial_path) - len(original_vehicle.path))
            
        diff_percentage = (diff_count / max(len(initial_path), len(original_vehicle.path))) * 100
        summary_data.append(["Path difference:", f"{diff_percentage:.1f}% of nodes are different"])
        
        # Compare initial and final travel times if available
        if iterations and iterations[0]['algorithm_results'] and iterations[-1]['algorithm_results']:
            for algo_name in iterations[0]['algorithm_results']:
                if algo_name in iterations[-1]['algorithm_results']:
                    initial_time = iterations[0]['algorithm_results'][algo_name]['travel_time']
                    final_time = iterations[-1]['algorithm_results'][algo_name]['travel_time']
                    change_pct = ((final_time - initial_time) / initial_time) * 100 if initial_time > 0 else 0
                    
                    summary_data.append([f"{algo_name} travel time change:", 
                                       f"{initial_time:.2f}s → {final_time:.2f}s ({change_pct:.1f}%)"])
    else:
        summary_data.append(["Path difference:", "No change in path"])
    
    for row in summary_data:
        ws_summary.append(row)
    
    # Format summary sheet
    ws_summary.merge_cells('A1:B1')
    ws_summary['A1'].font = Font(bold=True, size=14)
    
    # 2. Iterations Data sheet
    ws_iterations = wb.create_sheet("Iterations Data")
    
    # Create table with data from all iterations
    ws_iterations.append(["Stress Test Iterations Data"])
    ws_iterations.append([])
    
    # Create DataFrame for iteration data
    iteration_rows = []
    for iter_data in iterations:
        # Get algorithm data
        for algo_name, algo_results in iter_data['algorithm_results'].items():
            iteration_rows.append({
                'Iteration': iter_data['iteration'],
                'Vehicles Added': iter_data['vehicles_added'],
                'Total Vehicles': iter_data['total_vehicles'],
                'Mean Congestion': iter_data['mean_congestion'],
                'Algorithm': algo_name,
                'Travel Time (s)': algo_results['travel_time'],
                'Path Length': algo_results['path_length'],
                'Computation Time (s)': algo_results['computation_time']
            })
    
    # Convert to DataFrame
    df_iterations = pd.DataFrame(iteration_rows)
    
    # Write to Excel
    if not df_iterations.empty:
        for r in dataframe_to_rows(df_iterations, index=False, header=True):
            ws_iterations.append(r)
    else:
        ws_iterations.append(["No iteration data available"])
    
    # 3. Stress Test Analysis sheet
    ws_analysis = wb.create_sheet("Analysis Charts")
    
    ws_analysis.append(["Stress Test Analysis Charts"])
    ws_analysis.append([])
    
    # Create aggregated data for charts
    agg_data = []
    for iter_data in iterations:
        # Get best algorithm for this iteration
        best_algo = None
        best_time = float('inf')
        
        for algo_name, algo_results in iter_data['algorithm_results'].items():
            if algo_results['travel_time'] < best_time:
                best_time = algo_results['travel_time']
                best_algo = algo_name
        
        agg_data.append({
            'Iteration': iter_data['iteration'],
            'Total Vehicles': iter_data['total_vehicles'],
            'Mean Congestion': iter_data['mean_congestion'],
            'Best Algorithm': best_algo,
            'Best Travel Time': best_time
        })
    
    # Convert to DataFrame
    df_agg = pd.DataFrame(agg_data)
    
    # Write aggregated data
    ws_analysis.append(["Aggregated Data for Charts"])
    if not df_agg.empty:
        for r in dataframe_to_rows(df_agg, index=False, header=True):
            ws_analysis.append(r)
    
    # Add charts note
    ws_analysis.append([])
    ws_analysis.append(["Charts would be displayed here in the Excel file."])
    ws_analysis.append(["Please create charts manually using the data above."])
    
    # 4. Added Vehicles sheet
    ws_vehicles = wb.create_sheet("Added Vehicles")
    
    ws_vehicles.append(["Vehicles Added During Stress Test"])
    ws_vehicles.append([])
    
    # Create DataFrame for added vehicles
    vehicle_rows = []
    for v in added_vehicles:
        vehicle_rows.append({
            'Vehicle ID': v.id,
            'Source': v.source,
            'Destination': v.destination,
            'Path Length': len(v.path) if v.path else 0
        })
    
    # Convert to DataFrame
    df_vehicles = pd.DataFrame(vehicle_rows)
    
    # Write to Excel
    if not df_vehicles.empty:
        for r in dataframe_to_rows(df_vehicles, index=False, header=True):
            ws_vehicles.append(r)
    else:
        ws_vehicles.append(["No vehicles added during stress test"])
    
    # Save the workbook
    wb.save(excel_file)
    print(f"Iterative stress test analysis exported to {excel_file}")
    
    return excel_file


def export_stress_test_analysis(G, original_vehicle: Vehicle, original_path: List[int],
                               congestion_data: Dict[str, float], 
                               original_congestion: Dict[str, float],
                               added_vehicles: List[Vehicle]) -> str:
    """
    Export detailed analysis of a stress test, including before and after comparison.
    
    Args:
        G: Graph representing the road network
        original_vehicle: The vehicle that was stress tested
        original_path: The original path before stress testing
        congestion_data: Current congestion data (after stress test)
        original_congestion: Original congestion data (before stress test)
        added_vehicles: List of vehicles added during stress test
        
    Returns:
        Path to the created Excel file
    """
    print("Creating stress test analysis Excel file...")
    
    # Create output directory if it doesn't exist
    analysis_dir = os.path.join('london_simulation', 'stress_test_reports')
    os.makedirs(analysis_dir, exist_ok=True)
    
    # Create a timestamp for the filename
    timestamp = int(time.time())
    excel_file = os.path.join(analysis_dir, f"stress_test_vehicle_{original_vehicle.id}_{timestamp}.xlsx")
    
    # Create a new Excel workbook
    wb = Workbook()
    
    # Remove default worksheet
    ws0 = wb.active
    wb.remove(ws0)
    
    # 1. Summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    # Summary information
    summary_data = [
        ["Stress Test Analysis Report"],
        [""],
        ["Generated on:", time.strftime("%Y-%m-%d %H:%M:%S")],
        ["Vehicle ID:", original_vehicle.id],
        ["Source:", original_vehicle.source],
        ["Destination:", original_vehicle.destination],
        ["Stress test vehicles added:", len(added_vehicles)],
        [""],
        ["Path Comparison:"],
        ["Original path length:", len(original_path)],
        ["New path length:", len(original_vehicle.path)],
    ]
    
    # Calculate path difference
    if original_path != original_vehicle.path:
        diff_count = sum(1 for x, y in zip(original_path, original_vehicle.path) if x != y)
        if len(original_path) != len(original_vehicle.path):
            diff_count += abs(len(original_path) - len(original_vehicle.path))
            
        diff_percentage = (diff_count / max(len(original_path), len(original_vehicle.path))) * 100
        summary_data.append(["Path difference:", f"{diff_percentage:.1f}% of nodes are different"])
        
        # Compare travel times if available
        if 'A*' in original_vehicle.travel_times:
            summary_data.append(["Current travel time:", f"{original_vehicle.travel_times['A*']:.2f}s"])
    else:
        summary_data.append(["Path difference:", "No change in path"])
    
    for row in summary_data:
        ws_summary.append(row)
    
    # Format summary sheet
    ws_summary.merge_cells('A1:B1')
    ws_summary['A1'].font = Font(bold=True, size=14)
    
    # 2. Path Comparison sheet
    ws_path = wb.create_sheet("Path Comparison")
    
    # Create table comparing before and after paths
    ws_path.append(["Path Node Comparison: Before vs After Stress Test"])
    ws_path.append([])
    
    # Determine max path length
    max_length = max(len(original_path), len(original_vehicle.path))
    
    # Create headers
    ws_path.append(["Node Index", "Original Path Node", "New Path Node", "Changed?"])
    
    # Add path comparison
    for i in range(max_length):
        orig_node = original_path[i] if i < len(original_path) else "N/A"
        new_node = original_vehicle.path[i] if i < len(original_vehicle.path) else "N/A"
        changed = "Yes" if orig_node != new_node else "No"
        
        ws_path.append([i, orig_node, new_node, changed])
    
    # 3. Congestion Comparison sheet
    ws_congestion = wb.create_sheet("Congestion Comparison")
    
    # Create table comparing before and after congestion
    ws_congestion.append(["Congestion Comparison: Before vs After Stress Test"])
    ws_congestion.append([])
    
    # Create data for comparison
    congestion_rows = []
    
    # Focus on edges that are part of either path
    path_edges = set()
    
    # Get edges from original path
    for i in range(len(original_path) - 1):
        u, v = original_path[i], original_path[i+1]
        if u in G and v in G[u]:
            for k in G[u][v]:
                path_edges.add((u, v, k))
    
    # Get edges from new path
    for i in range(len(original_vehicle.path) - 1):
        u, v = original_vehicle.path[i], original_vehicle.path[i+1]
        if u in G and v in G[u]:
            for k in G[u][v]:
                path_edges.add((u, v, k))
    
    # Add data for each edge
    for u, v, k in path_edges:
        edge_id = f"{u}_{v}_{k}"
        
        # Get before and after congestion
        before_congestion = original_congestion.get(edge_id, 1.0)
        after_congestion = congestion_data.get(edge_id, 1.0)
        
        # Calculate change
        change = after_congestion - before_congestion
        percent_change = (change / before_congestion) * 100 if before_congestion > 0 else 0
        
        # Get vehicle count
        vehicle_count = G[u][v][k].get('vehicle_count', 0)
        
        # Get edge information
        length = G[u][v][k].get('length', 0)
        
        # Check if edge is in original path
        in_original = False
        for i in range(len(original_path) - 1):
            if original_path[i] == u and original_path[i+1] == v:
                in_original = True
                break
        
        # Check if edge is in new path
        in_new = False
        for i in range(len(original_vehicle.path) - 1):
            if original_vehicle.path[i] == u and original_vehicle.path[i+1] == v:
                in_new = True
                break
        
        # Get MM1 statistics if available
        mm1_stats = G[u][v][k].get('mm1_stats')
        if mm1_stats and mm1_stats.get('system_stable', False):
            utilization = mm1_stats.get('utilization', 'N/A')
            avg_vehicles = mm1_stats.get('avg_vehicles_in_system', 'N/A')
        else:
            utilization = "N/A"
            avg_vehicles = "N/A"
        
        congestion_rows.append({
            'Edge ID': edge_id,
            'Source': u,
            'Target': v,
            'Length (m)': length,
            'Before Congestion': before_congestion,
            'After Congestion': after_congestion,
            'Absolute Change': change,
            'Percent Change (%)': percent_change,
            'Vehicle Count': vehicle_count,
            'In Original Path': "Yes" if in_original else "No",
            'In New Path': "Yes" if in_new else "No",
            'MM1 Utilization': utilization,
            'MM1 Avg Vehicles': avg_vehicles
        })
    
    # Convert to DataFrame and sort by percent change
    df_congestion = pd.DataFrame(congestion_rows)
    if not df_congestion.empty:
        df_congestion = df_congestion.sort_values('Percent Change (%)', ascending=False)
        
        # Write to Excel
        for r in dataframe_to_rows(df_congestion, index=False, header=True):
            ws_congestion.append(r)
    else:
        ws_congestion.append(["No congestion data available"])
    
    # 4. Stress Test Vehicles sheet
    ws_vehicles = wb.create_sheet("Stress Test Vehicles")
    
    # Create table of added vehicles
    ws_vehicles.append(["Vehicles Added During Stress Test"])
    ws_vehicles.append([])
    
    vehicle_rows = []
    for v in added_vehicles:
        vehicle_rows.append({
            'Vehicle ID': v.id,
            'Source': v.source,
            'Destination': v.destination,
            'Path Length': len(v.path) if v.path else 0
        })
    
    # Convert to DataFrame
    df_vehicles = pd.DataFrame(vehicle_rows)
    if not df_vehicles.empty:
        # Write to Excel
        for r in dataframe_to_rows(df_vehicles, index=False, header=True):
            ws_vehicles.append(r)
    else:
        ws_vehicles.append(["No vehicles added during stress test"])
    
    # 5. Algorithm Comparison sheet (if multiple algorithms were used)
    if len(original_vehicle.paths) > 1:
        ws_algo = wb.create_sheet("Algorithm Comparison")
        
        ws_algo.append(["Algorithm Comparison After Stress Test"])
        ws_algo.append([])
        
        algo_rows = []
        for algo_name, path in original_vehicle.paths.items():
            if not path:
                continue
            
            travel_time = original_vehicle.travel_times.get(algo_name, 'N/A')
            comp_time = original_vehicle.computation_times.get(algo_name, 'N/A')
            
            algo_rows.append({
                'Algorithm': algo_name,
                'Path Length': len(path),
                'Travel Time (s)': travel_time,
                'Computation Time (s)': comp_time
            })
        
        # Convert to DataFrame and sort by travel time
        df_algo = pd.DataFrame(algo_rows)
        if not df_algo.empty:
            df_algo = df_algo.sort_values('Travel Time (s)')
            
            # Write to Excel
            for r in dataframe_to_rows(df_algo, index=False, header=True):
                ws_algo.append(r)
        else:
            ws_algo.append(["No algorithm comparison data available"])
    
    # Save the workbook
    wb.save(excel_file)
    print(f"Stress test analysis exported to {excel_file}")
    
    return excel_file