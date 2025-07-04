"""
Analysis and export module for the traffic simulation project.
Handles Excel reports, statistics, and data analysis.
"""

import pandas as pd
import numpy as np
import os
import time
import platform
import subprocess
from typing import Dict, List, Optional, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, LineChart, ScatterChart, Series

from models import Vehicle
from routing import create_congestion_graph


def open_excel_file(file_path: str) -> None:
    """
    Open an Excel file using the system's default application.
    
    Args:
        file_path: Path to the Excel file to open
    """
    if file_path and os.path.exists(file_path):
        try:
            if platform.system() == 'Darwin':       # macOS
                subprocess.call(('open', file_path))
            elif platform.system() == 'Windows':    # Windows
                os.startfile(file_path)
            else:                                   # Linux variants
                subprocess.call(('xdg-open', file_path))
            print(f"Opened Excel file: {file_path}")
        except Exception as e:
            print(f"Error opening Excel file: {e}")
    else:
        print("No Excel file available to open.")


def print_algorithm_comparison_table(G, vehicle: Vehicle, congestion_data: Dict[str, float]) -> pd.DataFrame:
    """
    Print a table comparing the performance of A*, Shortest Path, and Shortest Path Congestion Aware algorithms for a vehicle.
    
    Args:
        G: NetworkX graph
        vehicle: Vehicle to analyze
        congestion_data: Dictionary of congestion values
        
    Returns:
        DataFrame with comparison results
    """
    if not vehicle.paths:
        print("No paths calculated for this vehicle")
        return pd.DataFrame()
    
    # Create congestion graph to get accurate travel times
    G_congestion = create_congestion_graph(G, congestion_data)
    
    # Calculate travel times and distances for each path
    results = []
    
    for algo_name, path in vehicle.paths.items():
        if not path or len(path) < 2:
            continue
        
        if algo_name == 'Shortest Path':
            # Shortest Path now stores base travel time (no congestion)
            base_travel_time = vehicle.travel_times.get(algo_name, 0)
            
            # Calculate actual distance and travel time with congestion for comparison
            actual_travel_time = 0
            total_distance = 0
            total_congestion = 0
            max_congestion = 0
            
            for i in range(len(path) - 1):
                u, v = path[i], path[i+1]
                if u in G_congestion and v in G_congestion[u]:
                    actual_travel_time += G_congestion[u][v]['travel_time']
                    total_distance += G_congestion[u][v]['length']
                    congestion = G_congestion[u][v].get('congestion', 1.0)
                    total_congestion += congestion
                    max_congestion = max(max_congestion, congestion)
            
            avg_congestion = total_congestion / (len(path) - 1) if len(path) > 1 else 0
            
            results.append({
                'Algorithm': algo_name,
                'Path Length (nodes)': len(path),
                'Travel Time (s)': base_travel_time,  # Base time (no congestion)
                'Realistic Travel Time (s)': actual_travel_time,  # With congestion
                'Distance (m)': total_distance,
                'Computation Time (s)': vehicle.computation_times.get(algo_name, 0),
                'Service Rate (routes/sec)': vehicle.service_rates.get(algo_name, 0),
                'Avg Congestion': avg_congestion,
                'Max Congestion': max_congestion
            })
        else:
            # A* stores realistic travel time (with congestion)
            realistic_travel_time = vehicle.travel_times.get(algo_name, 0)
            
            # Calculate total distance and congestion
            total_distance = 0
            total_congestion = 0
            max_congestion = 0
            
            for i in range(len(path) - 1):
                u, v = path[i], path[i+1]
                if u in G_congestion and v in G_congestion[u]:
                    total_distance += G_congestion[u][v]['length']
                    congestion = G_congestion[u][v].get('congestion', 1.0)
                    total_congestion += congestion
                    max_congestion = max(max_congestion, congestion)
            
            avg_congestion = total_congestion / (len(path) - 1) if len(path) > 1 else 0
            
            results.append({
                'Algorithm': algo_name,
                'Path Length (nodes)': len(path),
                'Travel Time (s)': realistic_travel_time,
                'Distance (m)': total_distance,
                'Computation Time (s)': vehicle.computation_times.get(algo_name, 0),
                'Service Rate (routes/sec)': vehicle.service_rates.get(algo_name, 0),
                'Avg Congestion': avg_congestion,
                'Max Congestion': max_congestion
            })
    
    # Convert to DataFrame for easy display
    df = pd.DataFrame(results)
    
    # Sort by travel time
    df = df.sort_values('Travel Time (s)')
    
    # Print the table
    print("\nAlgorithm Comparison for Vehicle", vehicle.id)
    print(df.to_string(index=False))
    
    # Show best algorithm by different metrics
    if not df.empty:
        best_time = df.loc[df['Travel Time (s)'].idxmin(), 'Algorithm']
        best_service_rate = df.loc[df['Service Rate (routes/sec)'].idxmax(), 'Algorithm']
        print(f"\nBest travel time: {best_time}")
        print(f"Highest service rate: {best_service_rate}")
    
    return df


def simple_algorithm_comparison_table(G, vehicle: Vehicle, congestion_data: Dict[str, float]) -> pd.DataFrame:
    """
    Simplified comparison table that avoids complex edge data access.
    
    Args:
        G: NetworkX graph
        vehicle: Vehicle to analyze
        congestion_data: Dictionary of congestion values
        
    Returns:
        DataFrame with comparison results
    """
    if not vehicle.paths:
        print("No paths calculated for this vehicle")
        return pd.DataFrame()
    
    print(f"\nSimple Algorithm Comparison for Vehicle {vehicle.id}")
    print("=" * 60)
    
    results = []
    
    for algo_name, path in vehicle.paths.items():
        if not path:
            continue
            
        # Get basic metrics from vehicle object
        travel_time = vehicle.travel_times.get(algo_name, 0)
        comp_time = vehicle.computation_times.get(algo_name, 0)
        service_rate = vehicle.service_rates.get(algo_name, 0)
        
        results.append({
            'Algorithm': algo_name,
            'Path Length': len(path),
            'Travel Time (s)': travel_time,  # Both algorithms now store travel time
            'Computation Time': comp_time,
            'Service Rate': service_rate
        })
    
    # Convert to DataFrame
    df = pd.DataFrame(results)
    
    if not df.empty:
        print(df.to_string(index=False, float_format='%.4f'))
        
        # Show best by service rate
        best_service = df.loc[df['Service Rate'].idxmax(), 'Algorithm']
        print(f"\nHighest service rate: {best_service}")
        
        # Show optimization focus
        print(f"\nNote: A* and Shortest Path Congestion Aware consider congestion, Shortest Path ignores congestion (baseline)")
    
    return df


def print_travel_time_analysis_summary(G, vehicles: List[Vehicle], congestion_data: Dict[str, float]) -> None:
    """
    Analyze travel time penalties across all vehicles.
    
    Args:
        G: NetworkX graph
        vehicles: List of vehicles to analyze
        congestion_data: Dictionary of congestion values
    """
    print("\n=== Travel Time Analysis Summary ===")
    
    if not vehicles:
        print("No vehicles available for analysis")
        return
    
    # Create congestion graph
    G_congestion = create_congestion_graph(G, congestion_data)
    
    # Analyze all vehicles with calculated paths
    analyzed_vehicles = [v for v in vehicles if v.paths and 'A*' in v.paths]
    
    if not analyzed_vehicles:
        print("No vehicles with calculated A* paths available")
        return
    
    print(f"Analyzing {len(analyzed_vehicles)} vehicles with calculated routes:")
    
    total_base_time = 0
    total_actual_time = 0
    congestion_impacts = []
    
    for vehicle in analyzed_vehicles:
        if 'A*' not in vehicle.paths:
            continue
            
        path = vehicle.paths['A*']
        if len(path) < 2:
            continue
        
        # Calculate base vs actual travel time for this vehicle
        base_time = 0
        actual_time = vehicle.travel_times.get('A*', 0)
        
        for i in range(len(path) - 1):
            u, v = path[i], path[i+1]
            if u in G_congestion and v in G_congestion[u]:
                edge_keys = list(G_congestion[u][v].keys())
                if edge_keys:
                    edge_data = G_congestion[u][v][edge_keys[0]]
                    base_time += edge_data.get('base_travel_time', 0)
        
        if base_time > 0:
            congestion_impact = ((actual_time - base_time) / base_time) * 100
            congestion_impacts.append(congestion_impact)
            total_base_time += base_time
            total_actual_time += actual_time
    
    if congestion_impacts:
        avg_impact = sum(congestion_impacts) / len(congestion_impacts)
        max_impact = max(congestion_impacts)
        min_impact = min(congestion_impacts)
        
        overall_impact = ((total_actual_time - total_base_time) / total_base_time) * 100 if total_base_time > 0 else 0
        
        print(f"\nCongestion Impact on Travel Times:")
        print(f"  Average penalty: {avg_impact:.1f}%")
        print(f"  Maximum penalty: {max_impact:.1f}%")
        print(f"  Minimum penalty: {min_impact:.1f}%")
        print(f"  Overall network penalty: {overall_impact:.1f}%")
        
        # Categorize impact severity
        if overall_impact < 10:
            print(f"  📊 Network Status: Minimal congestion impact")
        elif overall_impact < 25:
            print(f"  📊 Network Status: Moderate congestion impact")
        elif overall_impact < 50:
            print(f"  📊 Network Status: High congestion impact")
        else:
            print(f"  📊 Network Status: Severe congestion impact")
            
        print(f"\nTotal time saved by fixing double penalization:")
        print(f"  Previous system would have added ~{overall_impact:.1f}% more penalties")
        print(f"  Current system uses realistic single penalties")
    else:
        print("No travel time data available for analysis")


def recalculate_routes_for_selected_vehicles(G, vehicles: List[Vehicle], 
                                           congestion_data: Dict[str, float], 
                                           scenario: str) -> List[Dict]:
    """
    Recalculate routes for all selected vehicles under current congestion.
    
    Args:
        G: NetworkX graph
        vehicles: List of vehicles
        congestion_data: Dictionary of congestion values
        scenario: Current scenario name
        
    Returns:
        List of summary dictionaries
    """
    from routing import calculate_all_routes
    from visualization import enhanced_visualize_congestion_map
    
    selected_vehicles = [v for v in vehicles if v.selected_for_analysis]
    
    if not selected_vehicles:
        print("No vehicles selected for analysis")
        return []
    
    print(f"\nRecalculating routes for {len(selected_vehicles)} selected vehicles under {scenario} scenario...")
    
    # Summary results
    summary = []
    
    for vehicle in selected_vehicles:
        print(f"\nVehicle {vehicle.id}: {vehicle.source} -> {vehicle.destination}")
        
        # Calculate new routes
        calculate_all_routes(G, vehicle, congestion_data)
        
        # Print comparison table
        try:
            result_df = print_algorithm_comparison_table(G, vehicle, congestion_data)
        except Exception as e:
            print(f"Warning: Complex comparison failed, using simple version: {e}")
            result_df = simple_algorithm_comparison_table(G, vehicle, congestion_data)
        
        # Add to summary
        if result_df is not None and not result_df.empty:
            best_time_row = result_df.loc[result_df['Travel Time (s)'].idxmin()]
            best_service_row = result_df.loc[result_df['Service Rate (routes/sec)'].idxmax()]
            
            summary.append({
                'Vehicle ID': vehicle.id,
                'Source': vehicle.source,
                'Destination': vehicle.destination,
                'Best Time Algorithm': best_time_row['Algorithm'],
                'Best Time (s)': best_time_row['Travel Time (s)'],
                'Highest Service Rate Algorithm': best_service_row['Algorithm'], 
                'Highest Service Rate': best_service_row['Service Rate (routes/sec)'],
                'Scenario': scenario
            })
    
    # Print summary table
    if summary:
        summary_df = pd.DataFrame(summary)
        print("\nSummary of Algorithm Performance for Selected Vehicles:")
        print(summary_df.to_string(index=False))
        
        # Count algorithm performance
        time_algo_counts = summary_df['Best Time Algorithm'].value_counts()
        service_algo_counts = summary_df['Highest Service Rate Algorithm'].value_counts()
        
        print("\nPerformance Summary:")
        print("Best Travel Time:")
        for algo, count in time_algo_counts.items():
            print(f"  {algo}: {count} out of {len(summary_df)} cases ({count/len(summary_df)*100:.1f}%)")
        
        print("Highest Service Rate:")
        for algo, count in service_algo_counts.items():
            print(f"  {algo}: {count} out of {len(summary_df)} cases ({count/len(summary_df)*100:.1f}%)")
    
    # Display the updated map with routes
    print("\nUpdated map showing all calculated routes:")
    enhanced_visualize_congestion_map(G, congestion_data, vehicles, scenario, None)
    
    return summary


def export_analysis_to_excel(G, vehicles: List[Vehicle], congestion_data: Dict[str, float], 
                           scenario: str, selected_vehicle: Optional[Vehicle] = None) -> str:
    """
    Export all analysis and calculations to an Excel file.
    
    Args:
        G: NetworkX graph
        vehicles: List of vehicles
        congestion_data: Dictionary of congestion values
        scenario: Current scenario name
        selected_vehicle: Specific vehicle to highlight in detailed analysis
        
    Returns:
        Path to the created Excel file
    """
    print("Creating comprehensive analysis Excel file...")
    
    # Create output directory if it doesn't exist
    analysis_dir = os.path.join('london_simulation', 'analysis_reports')
    os.makedirs(analysis_dir, exist_ok=True)
    
    # Create a timestamp for the filename
    timestamp = int(time.time())
    excel_file = os.path.join(analysis_dir, f"traffic_analysis_{scenario.replace(' ', '_')}_{timestamp}.xlsx")
    
    # Create a new Excel workbook
    wb = Workbook()
    
    # Remove default worksheet
    ws0 = wb.active
    wb.remove(ws0)
    
    # 1. Summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    # Summary information
    summary_data = [
        ["London Traffic Simulation Analysis Report"],
        [""],
        ["Generated on:", time.strftime("%Y-%m-%d %H:%M:%S")],
        ["Scenario:", scenario],
        ["Total vehicles:", len(vehicles)],
        ["Selected vehicles:", sum(1 for v in vehicles if v.selected_for_analysis)],
        [""],
        ["Congestion Statistics:"],
        ["Mean congestion:", np.mean(list(congestion_data.values()))],
        ["Min congestion:", np.min(list(congestion_data.values()))],
        ["Max congestion:", np.max(list(congestion_data.values()))],
        ["Standard deviation:", np.std(list(congestion_data.values()))],
    ]
    
    for row in summary_data:
        ws_summary.append(row)
    
    # Format summary sheet
    ws_summary.merge_cells('A1:B1')
    ws_summary['A1'].font = Font(bold=True, size=14)
    
    # 2. Vehicle List sheet
    ws_vehicles = wb.create_sheet("Vehicle List")
    
    # Create DataFrame for all vehicles
    vehicle_data = []
    for v in vehicles:
        # Get best algorithm if available
        best_algo = None
        best_time = float('inf')
        if v.travel_times:
            best_algo, best_time = min(v.travel_times.items(), key=lambda x: x[1])
        
        # Check if this vehicle has been analyzed with all algorithms
        algorithms_used = len(v.paths)
        
        vehicle_data.append({
            'Vehicle ID': v.id,
            'Source': v.source,
            'Destination': v.destination,
            'Path Length': len(v.path) if v.path else 0,
            'Selected for Analysis': "Yes" if v.selected_for_analysis else "No",
            'Best Algorithm': best_algo if best_algo else "N/A",
            'Best Travel Time': best_time if best_time != float('inf') else "N/A",
            'Algorithms Used': algorithms_used
        })
    
    # Convert to DataFrame
    df_vehicles = pd.DataFrame(vehicle_data)
    if not df_vehicles.empty:
        # Write to Excel
        for r in dataframe_to_rows(df_vehicles, index=False, header=True):
            ws_vehicles.append(r)
    else:
        ws_vehicles.append(["No vehicles in simulation"])
    
    # 3. Algorithm Comparison sheet
    ws_algo = wb.create_sheet("Algorithm Comparison")
    
    # Get data for vehicles that have been analyzed with multiple algorithms
    analyzed_vehicles = [v for v in vehicles if len(v.paths) > 1]
    
    if analyzed_vehicles:
        # First, create summary of which algorithm performs best most often
        algo_counts = {'Dijkstra': 0, 'A*': 0, 'Shortest Path': 0, 'Shortest Path Congestion Aware': 0, 'Cost-Efficient': 0}
        for v in analyzed_vehicles:
            if v.travel_times:
                best_algo = min(v.travel_times.items(), key=lambda x: x[1])[0]
                if best_algo in algo_counts:
                    algo_counts[best_algo] += 1
        
        ws_algo.append(["Algorithm Performance Summary"])
        ws_algo.append(["Algorithm", "Best Count", "Percentage"])
        
        for algo, count in algo_counts.items():
            if count > 0:
                percentage = (count / len(analyzed_vehicles)) * 100
                ws_algo.append([algo, count, f"{percentage:.1f}%"])
        
        ws_algo.append([])
        ws_algo.append(["Detailed Algorithm Comparison by Vehicle"])
        ws_algo.append([])
        
        # Then add detailed comparison for each analyzed vehicle
        for v in analyzed_vehicles:
            ws_algo.append([f"Vehicle {v.id}: {v.source} -> {v.destination}"])
            
            # Create comparison table for this vehicle
            algo_data = []
            for algo_name, path in v.paths.items():
                if not path:
                    continue
                
                algo_data.append({
                    'Algorithm': algo_name,
                    'Path Length': len(path),
                    'Travel Time (s)': v.travel_times.get(algo_name, "N/A"),
                    'Computation Time (s)': v.computation_times.get(algo_name, "N/A")
                })
            
            if algo_data:
                # Convert to DataFrame and sort by travel time
                df_algo = pd.DataFrame(algo_data)
                df_algo = df_algo.sort_values('Travel Time (s)')
                
                # Write to Excel
                for r in dataframe_to_rows(df_algo, index=False, header=True):
                    ws_algo.append(r)
                
                ws_algo.append([])
    else:
        ws_algo.append(["No vehicles have been analyzed with multiple algorithms"])
    
    # 4. Congestion Data sheet
    ws_congestion = wb.create_sheet("Congestion Data")
    
    # Create DataFrame for congestion data
    congestion_rows = []
    for u, v, k in G.edges(keys=True):
        edge_id = f"{u}_{v}_{k}"
        if edge_id in congestion_data:
            congestion_level = congestion_data[edge_id]
            vehicle_count = G[u][v][k].get('vehicle_count', 0)
            
            # Get MM1 statistics if available
            mm1_stats = G[u][v][k].get('mm1_stats')
            if mm1_stats and mm1_stats['system_stable']:
                utilization = mm1_stats['utilization']
                avg_vehicles = mm1_stats['avg_vehicles_in_system']
                avg_time = mm1_stats['avg_time_in_system']
            else:
                utilization = "N/A"
                avg_vehicles = "N/A"
                avg_time = "N/A"
            
            congestion_rows.append({
                'Edge ID': edge_id,
                'Source': u,
                'Target': v,
                'Key': k,
                'Congestion Level': congestion_level,
                'Vehicle Count': vehicle_count,
                'Length (m)': G[u][v][k].get('length', 0),
                'MM1 Utilization': utilization,
                'MM1 Avg Vehicles': avg_vehicles,
                'MM1 Avg Time': avg_time
            })
    
    # Convert to DataFrame and sort by congestion level
    df_congestion = pd.DataFrame(congestion_rows)
    if not df_congestion.empty:
        df_congestion = df_congestion.sort_values('Congestion Level', ascending=False)
        
        # Write to Excel
        for r in dataframe_to_rows(df_congestion, index=False, header=True):
            ws_congestion.append(r)
    else:
        ws_congestion.append(["No congestion data available"])
    
    # 5. Detailed Vehicle Analysis sheet (if a specific vehicle is selected)
    if selected_vehicle:
        ws_detail = wb.create_sheet(f"Vehicle {selected_vehicle.id} Detail")
        
        ws_detail.append([f"Detailed Analysis for Vehicle {selected_vehicle.id}"])
        ws_detail.append(["Source:", selected_vehicle.source])
        ws_detail.append(["Destination:", selected_vehicle.destination])
        ws_detail.append([])
        
        # Path details
        if selected_vehicle.path:
            ws_detail.append(["Current Path:"])
            ws_detail.append(["Node", "Coordinates (x,y)", "Congestion"])
            
            for i, node in enumerate(selected_vehicle.path):
                x = G.nodes[node].get('x', 'N/A')
                y = G.nodes[node].get('y', 'N/A')
                
                # Get average congestion of edges leading to this node
                congestion = "N/A"
                if i > 0:
                    prev_node = selected_vehicle.path[i-1]
                    edge_id = f"{prev_node}_{node}_0"  # Assuming key=0 for simplicity
                    congestion = congestion_data.get(edge_id, "N/A")
                
                ws_detail.append([node, f"({x}, {y})", congestion])
            
            ws_detail.append([])
        
        # Algorithm comparison
        if selected_vehicle.paths:
            ws_detail.append(["Algorithm Comparison:"])
            
            algo_data = []
            for algo_name, path in selected_vehicle.paths.items():
                if not path:
                    continue
                
                algo_data.append({
                    'Algorithm': algo_name,
                    'Path Length': len(path),
                    'Travel Time (s)': selected_vehicle.travel_times.get(algo_name, "N/A"),
                    'Computation Time (s)': selected_vehicle.computation_times.get(algo_name, "N/A"),
                    'Service Rate (routes/sec)': selected_vehicle.service_rates.get(algo_name, "N/A")
                })
            
            if algo_data:
                # Convert to DataFrame and sort by travel time
                df_algo = pd.DataFrame(algo_data)
                df_algo = df_algo.sort_values('Travel Time (s)')
                
                # Write to Excel
                for r in dataframe_to_rows(df_algo, index=False, header=True):
                    ws_detail.append(r)
    
    # Save the workbook
    wb.save(excel_file)
    print(f"Analysis exported to {excel_file}")
    
    return excel_file


def export_impact_report_to_excel(G, vehicles: List[Vehicle], congestion_data: Dict[str, float],
                                original_congestion: Dict[str, float], scenario: str,
                                impact_report: Dict) -> str:
    """
    Export the vehicle impact report to an Excel file.
    
    Args:
        G: NetworkX graph
        vehicles: List of vehicles
        congestion_data: Current congestion data
        original_congestion: Original congestion data
        scenario: Current scenario name
        impact_report: Impact report dictionary
        
    Returns:
        Path to the created Excel file
    """
    # Create output directory
    reports_dir = os.path.join('london_simulation', 'impact_reports')
    os.makedirs(reports_dir, exist_ok=True)
    
    # Create a DataFrame for the summary
    summary_data = {
        'Metric': [
            'Total Vehicles',
            'Roads Affected (%)',
            'Average Vehicles per Road',
            'Overall Congestion Increase (%)',
            'Travel Time Increase (%)'
        ],
        'Value': [
            impact_report['vehicle_count'],
            impact_report['affected_pct'],
            impact_report['avg_vehicle_count'],
            impact_report['congestion_increase_pct'],
            impact_report['avg_time_increase_pct']
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    
    # Create a DataFrame for road type impact
    road_type_data = {
        'Road Type': ['Small Roads (<=100m)', 'Medium Roads (100-300m)', 'Major Roads (>300m)'],
        'Average Vehicles': [
            impact_report['avg_vehicles_short_roads'],
            impact_report['avg_vehicles_medium_roads'],
            impact_report['avg_vehicles_long_roads']
        ]
    }
    road_type_df = pd.DataFrame(road_type_data)
    
    # Create a DataFrame for travel time impacts
    if impact_report['travel_time_impacts']:
        time_impact_data = {
            'Vehicle ID': [impact['vehicle_id'] for impact in impact_report['travel_time_impacts']],
            'Base Travel Time (s)': [impact['base_time'] for impact in impact_report['travel_time_impacts']],
            'Current Travel Time (s)': [impact['current_time'] for impact in impact_report['travel_time_impacts']],
            'Increase (%)': [impact['increase_pct'] for impact in impact_report['travel_time_impacts']]
        }
        time_impact_df = pd.DataFrame(time_impact_data)
    else:
        time_impact_df = pd.DataFrame({'No Data': []})
    
    # Create a DataFrame for edge congestion
    edge_data = []
    for u, v, k in G.edges(keys=True):
        edge_id = f"{u}_{v}_{k}"
        if edge_id in congestion_data and edge_id in original_congestion:
            vehicle_count = G[u][v][k].get('vehicle_count', 0)
            base_cong = original_congestion[edge_id]
            current_cong = congestion_data[edge_id]
            
            edge_data.append({
                'Edge ID': edge_id,
                'Length (m)': G[u][v][k].get('length', 0),
                'Vehicle Count': vehicle_count,
                'Base Congestion': base_cong,
                'Current Congestion': current_cong,
                'Congestion Increase (%)': ((current_cong - base_cong) / base_cong * 100) if base_cong > 0 else 0
            })
    
    # Sort by vehicle count in descending order
    edge_df = pd.DataFrame(edge_data)
    if not edge_df.empty:
        edge_df = edge_df.sort_values('Vehicle Count', ascending=False)
    
    # Create the Excel file
    timestamp = int(time.time())
    excel_file = os.path.join(reports_dir, f"impact_report_{scenario.replace(' ', '_')}_{timestamp}.xlsx")
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        road_type_df.to_excel(writer, sheet_name='Road Types', index=False)
        time_impact_df.to_excel(writer, sheet_name='Travel Times', index=False)
        if not edge_df.empty:
            edge_df.to_excel(writer, sheet_name='Edge Data', index=False)
    
    print(f"  Saved impact report to {excel_file}")
    return excel_file